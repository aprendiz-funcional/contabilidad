import openpyxl, os, json, random, re
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl import load_workbook, Workbook
from openpyxl.styles import numbers

from dotenv import load_dotenv

class   ModelarInformacion():
    # Metodo constructor
    def __init__(self):
        load_dotenv()
        path_ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.path_documentos_ = os.path.join(path_,"Documentos")
        self.path_documentos = os.path.join(path_,"Documentos", "DIVIDENDOS.xlsx")
        self.path_procesados = os.path.join(path_,"Procesados")
        #Ruta del Json
        self.__path_json = os.path.join(path_,'modules','Json','columnas_necesarias.json')
    # Metodo principal
    def main(self):
        try:
            reteica = os.getenv("reteica")
            retefuente = os.getenv("retefuente")
            pago_accionistas = os.getenv("pago_accionistas")
            self.__crear_carpeta()
            columnas_retencion, mensaje = self.__importar_json('retencion')
            print(columnas_retencion)
            # Obtener los datos depurados
            df_datos = self.__depurar_informacion_dividendos()
            if df_datos is None:
                return "No se pudieron depurar los datos."
            
            # 📌--------------------Extraer datos de retefuente----------------------------------
            df_retefuente = df_datos[df_datos["CUENTA"] == retefuente][columnas_retencion].copy()
            # 📌 Agrupar por TERCERO y sumar los valores de ReteFuente
            df_retefuente = df_retefuente.groupby(["TERCERO", "Nombre Tercero"], as_index=False).sum()
            # 📌 Renombrar la columna CREDITOS a ReteFuente
            df_retefuente.rename(columns={"CREDITOS": "ReteFuente"}, inplace=True)
            # 📌 Verificar si hay datos
            if df_retefuente.empty:
                print("📌 No se encontraron datos de retefuente.")
                return "No se encontraron datos de retefuente."
            self.__guardar_datos(df_retefuente, 'rete_fuente')

            # 📌------------------ Extraer datos de reteica-------------------------------
            df_reteica = df_datos[df_datos["CUENTA"] == reteica][columnas_retencion].copy()
            # 📌 Agrupar por TERCERO y sumar los valores de ReteIca
            df_reteica = df_reteica.groupby(["TERCERO", "Nombre Tercero"], as_index=False).sum()
            # 📌 Renombrar la columna CREDITOS a ReteIca
            df_reteica.rename(columns={"CREDITOS": "ReteIca"}, inplace=True)
            # 📌 Verificar si hay datos
            if df_reteica.empty:
                print("📌 No se encontraron datos de reteica.")
                return "No se encontraron datos de reteica."
            self.__guardar_datos(df_reteica, 'rete_ica')

            # 📌------------------------- Extraer datos de pagos a accionistas-------------------------------------
            df_pagos_accionistas = self.__extraer_pagos_accionistas(df_datos, pago_accionistas, 'pago_accionistas')
            if isinstance(df_pagos_accionistas, str):
                return df_pagos_accionistas
            # 📌 -------------------Extraer datos de C x P contabilidad-------------------------------
            df_c_xp_contabilidad, mensaje = self.__leer_datos_excel(self.path_documentos, ' c xp contabilidad')
            if df_c_xp_contabilidad is None:            
                return "No se pudieron leer los datos de C x P contabilidad."
            # 📌 -------------------------------------Cruzar los datos----------------------------------
            df_datos_cruzados = self.__cruzar_datos(df_pagos_accionistas, df_retefuente, df_reteica, df_c_xp_contabilidad)
            # 📌 Verificar si hay datos
            if isinstance(df_datos_cruzados, str):
                return df_datos_cruzados
            # 📌 Aplicar estilos al archivo de Excel
            self.__aplicar_estilos_excel(self.path_documentos)
            # Nombre original y nuevo nombre del archivo
            año = datetime.now().year
            nuevo_nombre = os.path.join(self.path_documentos_, f"DIVIDENDOS PROCESADOS {año}.xlsx")

            # Renombrar el archivo
            os.rename(self.path_documentos, nuevo_nombre)

            print(f"Archivo renombrado a: {nuevo_nombre}")
            return df_datos_cruzados
        except Exception as e:
            mensaje = f"ERROR al extraer retefuente: {e}"
            print(mensaje)
            return mensaje
    #Metodo para cruzar los datos
    def __cruzar_datos(self, df_pagos_accionistas, df_retefuente, df_reteica, df_c_xp_contabilidad):
        try:
            print(f"📌 Datos de pagos a accionistas antes del cruce: {len(df_pagos_accionistas)} registros")
            
            # 📌 Unir df_pagos_accionistas con df_retefuente por "TERCERO"
            df_final = df_pagos_accionistas.merge(
                df_retefuente, on=["TERCERO", "Nombre Tercero"], how="left"
            )
            
            # 📌 Unir con df_reteica por "TERCERO"
            df_final = df_final.merge(
                df_reteica, on=["TERCERO", "Nombre Tercero"], how="left"
            )
            
            # 📌 Seleccionar solo las columnas necesarias de df_c_xp_contabilidad para evitar duplicados
            df_c_xp_contabilidad = df_c_xp_contabilidad[["TERCERO", "Nombre Tercero", "Saldo Final"]].drop_duplicates()
            
            print(f"📌 Datos de contabilidad antes del cruce: {len(df_c_xp_contabilidad)} registros")
            
            # 📌 Unir con df_c_xp_contabilidad por "TERCERO"
            df_final = df_final.merge(
                df_c_xp_contabilidad, on=["TERCERO", "Nombre Tercero"], how="left"
            )
            print(df_final)
            print(f"📌 Datos después del cruce: {len(df_final)} registros")
            
            # 📌 Asegurar que "Saldo Final" conserve su valor original si ya existía en df_pagos_accionistas
            if "Saldo Final" in df_pagos_accionistas.columns:
                df_final["Saldo Final"] = df_final["Saldo Final"].combine_first(df_pagos_accionistas["Saldo Final"])
            
            # 📌 Filtrar columnas finales según configuración
            columnas_final, mensaje = self.__importar_json('columnas_final')
            columnas_final = [col for col in columnas_final if col in df_final.columns]
            df_final = df_final[columnas_final]
            
            # 📌 Guardar los datos finales
            self.__guardar_datos(df_final, 'Cruce_final')
            
            return df_final
        
        except Exception as e:
            mensaje = f"❌ ERROR al cruzar los datos: {e}"
            print(mensaje)
            return mensaje

    def __cruzar_datos1(self, df_pagos_accionistas, df_retefuente, df_reteica, df_c_xp_contabilidad):
        try:
            print(f"📌 Datos de pagos a accionistas antes del cruce: {len(df_pagos_accionistas)} registros")
            
            # 📌 Unir df_pagos_accionistas con df_retefuente por "TERCERO"
            df_final = df_pagos_accionistas.merge(
                df_retefuente, on=["TERCERO", "Nombre Tercero"], how="left"
            )
            
            # 📌 Unir con df_reteica por "TERCERO"
            df_final = df_final.merge(
                df_reteica, on=["TERCERO", "Nombre Tercero"], how="left"
            )
            
            # 📌 Seleccionar solo las columnas necesarias de df_c_xp_contabilidad para evitar duplicados
            df_c_xp_contabilidad = df_c_xp_contabilidad[["TERCERO", "Nombre Tercero", "Saldo Final"]]
            
            print(f"📌 Datos de contabilidad antes del cruce: {len(df_c_xp_contabilidad)} registros")
            
            # 📌 Unir con df_c_xp_contabilidad por "TERCERO"
            df_final = df_final.merge(
                df_c_xp_contabilidad, on=["TERCERO", "Nombre Tercero"], how="left"
            )
            
            print(f"📌 Datos después del cruce: {len(df_final)} registros")
            
            # 📌 Asegurar que "Saldo Final" no sea reemplazado por NaN
            if "Saldo Final" in df_final.columns:
                df_final["Saldo Final"].fillna(0, inplace=True)
            
            # 📌 Filtrar columnas finales según configuración
            columnas_final, mensaje = self.__importar_json('columnas_final')
            columnas_final = [col for col in columnas_final if col in df_final.columns]
            df_final = df_final[columnas_final]
            
            # 📌 Guardar los datos finales
            self.__guardar_datos(df_final, 'Cruce_final')
            
            return df_final
        
        except Exception as e:
            mensaje = f"❌ ERROR al cruzar los datos: {e}"
            print(mensaje)
            return mensaje

    def __cruzar_datos1(self, df_pagos_accionistas, df_retefuente, df_reteica, df_c_xp_contabilidad):
        try:
            saldos = df_c_xp_contabilidad["Saldo Final"]

            print(f"📌 Datos de pagos a accionistas: {saldos}"      )
            # 📌 Unir df_pagos_accionistas con df_retefuente por "TERCERO"
            df_final = df_pagos_accionistas.merge(
                df_retefuente, on=["TERCERO", "Nombre Tercero"], how="left"
            )

            # 📌 Unir con df_reteica por "TERCERO"
            df_final = df_final.merge(
                df_reteica, on=["TERCERO", "Nombre Tercero"], how="left"
            )

            # 📌 Unir con df_c_xp_contabilidad por "TERCERO"
            df_c_xp_contabilidad = df_c_xp_contabilidad[["TERCERO", "Nombre Tercero", "Saldo Final"]]
            print(f"📌 Datos de pagos a accionistas: {df_c_xp_contabilidad}"     )
            df_final = df_final.merge(
                df_c_xp_contabilidad, on=["TERCERO", "Nombre Tercero"], how="left"
            )

            # 📌 Rellenar valores NaN con 0 para evitar errores en cálculos posteriores
            df_final.fillna(0, inplace=True)
            columnas_final, mensaje = self.__importar_json('columnas_final')
            columnas_final = [col for col in columnas_final if col in df_final.columns]
            df_final = df_final[columnas_final]
            self.__guardar_datos(df_final, 'Cruce_final')
            return df_final

        except Exception as e:
            mensaje = f"❌ ERROR al cruzar los datos: {e}"
            print(mensaje)
            return mensaje
    #Metodo para extraer los datos del pago de accionistas
    def __extraer_pagos_accionistas(self, df_datos, cuenta, nom_hoja):
        try:
            # 📌 Filtrar datos por la cuenta específica
            df_filtrado = df_datos[df_datos["CUENTA"] == cuenta].copy()
            if df_filtrado.empty:
                mensaje = f"⚠️ No hay datos para la cuenta {cuenta}"
                print(mensaje)
                return mensaje

            # 📌 Asegurar que 'TERCERO' se maneje como string con formato adecuado
            df_filtrado["TERCERO"] = df_filtrado["TERCERO"].astype(str).str.strip()
            df_filtrado["TERCERO"] = df_filtrado["TERCERO"].apply(lambda x: "{:,.0f}".format(int(x)) if x.isdigit() else x)

            # 📌 Convertir 'Nombre Mes' a minúsculas para estandarización
            df_filtrado["Nombre Mes"] = df_filtrado["Nombre Mes"].str.lower()

            # 📌 Agrupar datos para evitar duplicados, sumando valores de 'DEBITOS'
            df_agrupado = df_filtrado.groupby(["TERCERO", "Nombre Tercero", "Nombre Mes"], as_index=False)["DEBITOS"].sum()

            # 📌 Reestructurar los datos: convertir cada mes en una columna
            df_pivot = df_agrupado.pivot_table(
                index=["TERCERO", "Nombre Tercero"], 
                columns="Nombre Mes", 
                values="DEBITOS", 
                aggfunc="sum", 
                fill_value=0
            ).reset_index()

            # 📌 Lista completa de meses
            meses_completos = ["enero", "febrero", "marzo", "abril", "mayo", "junio", 
                                "julio", "agosto", "septiembre", 
                                "octubre", "noviembre", "diciembre"]

            # 📌 Agregar cualquier mes faltante con valores 0
            for mes in meses_completos:
                if mes not in df_pivot.columns:
                    df_pivot[mes] = 0

            # 📌 Definir bloques de meses con sus respectivas sumas
            meses_bloques = [
                (["enero", "febrero", "marzo", "abril", "mayo", "junio"], "Suma Enero-Junio"),
                (["julio", "agosto", "septiembre"], "Suma Julio-Septiembre"),
                (["octubre", "noviembre", "diciembre"], "Suma Octubre-Diciembre")
            ]

            # 📌 Insertar sumas después de cada bloque de meses
            for meses, suma_col in meses_bloques:
                df_pivot[suma_col] = df_pivot[meses].sum(axis=1)
            # 📌 Agregar columnas adicionales con valores fijos
            df_pivot["COMPROBANTE"] = "250"
            df_pivot["SECUENCIA"] = "350"
            df_pivot["FUENTE"] = "450"
            # 📌 Ordenar columnas dinámicamente con sumas después de cada bloque
            columnas_ordenadas = ["TERCERO", "Nombre Tercero", "COMPROBANTE","SECUENCIA", "FUENTE"]
            for meses, suma_col in meses_bloques:
                columnas_ordenadas.extend(meses + [suma_col])

            df_pivot = df_pivot[columnas_ordenadas]

            # 📌 Guardar datos procesados
            self.__guardar_datos(df_pivot, nom_hoja)
            return df_pivot

        except Exception as e:
            mensaje = f"❌ ERROR al extraer pagos accionistas: {e}"
            print(mensaje)
            return mensaje

    def __extraer_pagos_accionistas1(self, df_datos, cuenta, nom_hoja):
        try:
            # 📌 Filtrar datos por la cuenta específica
            df_filtrado = df_datos[df_datos["CUENTA"] == cuenta].copy()
            if df_filtrado.empty:
                mensaje = f"⚠️ No hay datos para la cuenta {cuenta}"
                print(mensaje)
                return mensaje

            # 📌 Asegurar que 'TERCERO' se maneje como string con formato adecuado
            df_filtrado["TERCERO"] = df_filtrado["TERCERO"].astype(str).str.strip()
            df_filtrado["TERCERO"] = df_filtrado["TERCERO"].apply(lambda x: "{:,.0f}".format(int(x)) if x.isdigit() else x)

            # 📌 Convertir 'Nombre Mes' a minúsculas para estandarización
            df_filtrado["Nombre Mes"] = df_filtrado["Nombre Mes"].str.lower()

            # 📌 Reestructurar los datos: convertir cada mes en una columna
            df_pivot = df_filtrado.pivot_table(
                index=["TERCERO", "Nombre Tercero", "COMPROBANTE", "SECUENCIA", "FUENTE"], 
                columns="Nombre Mes", 
                values="DEBITOS", 
                aggfunc="sum", 
                fill_value=0
            ).reset_index()

            # 📌 Lista completa de meses
            meses_completos = ["enero", "febrero", "marzo", "abril", "mayo", "junio", 
                            "julio", "agosto", "septiembre", 
                            "octubre", "noviembre", "diciembre"]

            # 📌 Agregar cualquier mes faltante con valores 0
            for mes in meses_completos:
                if mes not in df_pivot.columns:
                    df_pivot[mes] = 0

            # 📌 Definir bloques de meses con sus respectivas sumas
            meses_bloques = [
                (["enero", "febrero", "marzo", "abril", "mayo", "junio"], "Suma Enero-Junio"),
                (["julio", "agosto", "septiembre"], "Suma Julio-Septiembre"),
                (["octubre", "noviembre", "diciembre"], "Suma Octubre-Diciembre")
            ]

            # 📌 Insertar sumas después de cada bloque de meses
            for meses, suma_col in meses_bloques:
                df_pivot[suma_col] = df_pivot[meses].sum(axis=1)

            # 📌 Ordenar columnas dinámicamente con sumas después de cada bloque
            columnas_ordenadas = ["TERCERO", "Nombre Tercero", "COMPROBANTE", "SECUENCIA", "FUENTE"]
            for meses, suma_col in meses_bloques:
                columnas_ordenadas.extend(meses + [suma_col])

            df_pivot = df_pivot[columnas_ordenadas]

            # 📌 Guardar datos procesados
            self.__guardar_datos(df_pivot, nom_hoja)
            return df_pivot

        except Exception as e:
            mensaje = f"❌ ERROR al extraer pagos accionistas: {e}"
            print(mensaje)
            return mensaje
    #Metodo para aplicar estilos a los archivos de excel  
    def __depurar_informacion_dividendos(self):
        try:
            df_datos, mensaje = self.__leer_datos_excel(self.path_documentos, 'DIVIDENDOS')

            if df_datos is None:
                print(f"📌 ERROR: {mensaje}")
                return mensaje
            # Cargar columnas desde JSON
            columnas_permitidas, _ = self.__importar_json('columnas_permitidas')
            #columnas_monedas, _ = self.__importar_json('columnas_monedas')
            #columnas_texto, _ = self.__importar_json('columnas_texto')

            if not isinstance(columnas_permitidas, list):
                print("📌 ERROR: El JSON de columnas no tiene un formato válido.")
                return "Error en configuración de columnas"
            # Normalizar nombres de columnas
            df_datos.columns = df_datos.columns.str.strip()
            # Filtrar solo las columnas necesarias
            columnas_validas = [col for col in columnas_permitidas if col in df_datos.columns]
            df_filtrado = df_datos[columnas_validas].copy()
            """ # 🔹 Convertir columnas de moneda a float (manejando símbolos y separadores)
            for col in columnas_monedas:
                if col in df_filtrado.columns:
                    # Asegurar que todos los valores sean string
                    df_filtrado[col] = df_filtrado[col].astype(str).str.strip()
                    # 🔹 Reemplazar símbolos de moneda (€, $, ₡, etc.) y espacios
                    df_filtrado[col] = df_filtrado[col].apply(lambda x: re.sub(r'[^\d,.-]', '', x) if isinstance(x, str) else x)
                    # 🔹 Manejar formato europeo "1.234,56" → "1234.56"
                    df_filtrado[col] = df_filtrado[col].apply(lambda x: x.replace('.', '') if x.count(',') == 1 else x)
                    df_filtrado[col] = df_filtrado[col].apply(lambda x: x.replace(',', '.') if x.count(',') == 1 else x)
                    # 🔹 Convertir a float
                    df_filtrado[col] = pd.to_numeric(df_filtrado[col], errors='coerce')
            # 🔹 Mantener columnas de identificación como texto
            for col in columnas_texto:
                if col in df_filtrado.columns:
                    df_filtrado[col] = df_filtrado[col].astype(str).str.strip()  # Asegurar que sea string

                    # 🔹 Verificar si el valor es numérico y convertirlo con separadores de miles
                    df_filtrado[col] = df_filtrado[col].apply(lambda x: "{:,.0f}".format(int(x)) if x.isdigit() else x)

                    # 🔹 Reemplazar comas por puntos (si es necesario)
                    df_filtrado[col] = df_filtrado[col].str.replace(',', '.') """

            self.__guardar_datos(df_filtrado, 'DIVIDENDOS_DEPURADOS')   
            return df_filtrado

        except Exception as e:
            print(f"📌 ERROR al depurar información: {e}")
            return None
    # Metodo para guardar los datos en un archivo de excel
    def __guardar_datos(self, df_datos, nomhoja):
        try:
            #df_datos = self.__formatiar_columnas(df_datos)
            if df_datos is None or df_datos.empty:
                print("📌 ERROR: No hay datos para guardar.")
                return "No hay datos"
            columnas_texto, _ = self.__importar_json('columnas_texto')
            for col in columnas_texto:
                if col in df_datos.columns:
                    df_datos[col] = df_datos[col].astype(str).str.strip()  # Asegurar que sea string

                    # 🔹 Verificar si el valor es numérico y convertirlo con separadores de miles
                    df_datos[col] = df_datos[col].apply(lambda x: "{:,.0f}".format(int(x)) if x.isdigit() else x)

                    # 🔹 Reemplazar comas por puntos (si es necesario)
                    df_datos[col] = df_datos[col].str.replace(',', '.')
            # Guardar en el mismo archivo sin sobrescribir otras hojas
            with pd.ExcelWriter(self.path_documentos, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
                df_datos.to_excel(writer, sheet_name=nomhoja, index=False)

            # 🔹 Aplicar formato en Excel
            wb = load_workbook(self.path_documentos)
            ws = wb[nomhoja]
            columnas_monedas, _ = self.__importar_json('columnas_monedas')
            

            # Aplicar formato de moneda
            for col in columnas_monedas:
                if col in df_datos.columns:
                    col_idx = df_datos.columns.get_loc(col) + 1
                    for row in range(2, len(df_datos) + 2):
                        ws.cell(row=row, column=col_idx).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE  # $1,000.00

            """ # 🔹 Aplicar formato de TEXTO en la columna de identificación
            for col in columnas_texto:
                if col in df_datos.columns:
                    col_idx = df_datos.columns.get_loc(col) + 1
                    for row in range(2, len(df_datos) + 2):
                        cell = ws.cell(row=row, column=col_idx)
                        
                        # 📌 Aplicar formato de número con separador de miles
                        cell.number_format = "#,##0" """

            wb.save(self.path_documentos)
            wb.close()

            print(f"📌 Datos guardados correctamente en '{nomhoja}'.")
            return f"Datos guardados en '{nomhoja}'"

        except Exception as e:
            print(f"📌 ERROR al guardar los datos: {e}")
            return str(e)
    # Metodo para leer los datos de un archivo de excel
    def __leer_datos_excel(self, ruta, nomhoja):
        try:
            # Cargar el archivo de Excel
            wb = openpyxl.load_workbook(ruta, data_only=True)
            ws = wb[nomhoja]
            # Convertir los datos de la hoja a un DataFrame
            datos = ws.values
            columnas = next(datos)  # Extraer la primera fila como encabezados
            df = pd.DataFrame(datos, columns=columnas)  # Crear el DataFrame
            # Cerrar el archivo
            wb.close()
            mensaje = f"Datos de Excel leídos con éxito en la hoja '{nomhoja}'."
            # Eliminar vacios que esten en la columna identificación
            df.dropna(subset=['TERCERO'], inplace=True)
            return df, mensaje
        except Exception as e:
            mensaje = f"ERROR al leer los datos de Excel: {e}"
            print(mensaje)
            return None, mensaje
    #Metodo para crear las carpetas
    def __crear_carpeta(self):
            try:
                if not os.path.exists(self.path_procesados):
                    os.mkdir(self.path_procesados)
                Certificados = os.path.join(self.path_procesados, 'Certificados')
                if not os.path.exists(Certificados):
                    os.mkdir(Certificados)
                Certificados_excel = os.path.join(Certificados, 'Certificados_excel')
                if not os.path.exists(Certificados_excel):
                    os.mkdir(Certificados_excel)
                Certificados_pdf = os.path.join(Certificados, 'Certificados_pdf')
                if not os.path.exists(Certificados_pdf):
                    os.mkdir(Certificados_pdf)
                Certificados_word = os.path.join(Certificados, 'Certificados_word')
                if not os.path.exists(Certificados_word):
                    os.mkdir(Certificados_word)
            except Exception as e:
                self.logger.log(f"ERROR creando las carpetas: {e}<br>")
    #Metodo para tomar el diccionario del json
    def __importar_json(self, nom_json):
        try:                    
            with open(self.__path_json, "r", encoding="utf-8") as file:
                users = json.load(file)
            # Acceder a los datos
            campos_JSON = users.get(nom_json, {})
            mensaje = f"Json importado correctamente {nom_json}"
            print(mensaje)
            return campos_JSON, mensaje
        except Exception as e:
            mensaje = f"E importando el Json: del formulario {nom_json}"
            print(f"{mensaje} {e}")
            return None, mensaje
    # Metodo para aplicar estilos a los archivos de excel
    def __aplicar_estilos_excel(self, ruta_excel):
        try:
            # Cargar el archivo existente
            wb = load_workbook(ruta_excel)
            # Definir estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            alignment_encabezado = Alignment(horizontal="center", vertical="center")
            alignment_cuerpo = Alignment(horizontal="left", vertical="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            # Generar colores para las pestañas de las hojas
            def generar_color_aleatorio():
                return f"{random.randint(0, 255):02X}{random.randint(0, 255):02X}{random.randint(0, 255):02X}"
            # Iterar por cada hoja del archivo
            for ws in wb.worksheets:
                # Aplicar estilos a los encabezados (primera fila)
                for cell in ws[1]:  # Supone que la primera fila son los encabezados
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment_encabezado
                    cell.border = thin_border
                # Aplicar bordes y alineación a todas las celdas con datos
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.alignment = alignment_cuerpo
                        cell.border = thin_border
                # Ajustar el ancho de las columnas automáticamente
                for col in ws.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = max_length + 2
                # Asignar color aleatorio a la pestaña de la hoja
                ws.sheet_properties.tabColor = generar_color_aleatorio()
            # Guardar los cambios
            wb.save(ruta_excel)
            print("Estilos aplicados y colores asignados correctamente a las hojas del archivo de Excel.")
        except Exception as e:
            print(f"ERROR: No se pudieron aplicar estilos y colores al archivo de Excel: {str(e)}")
    # Metodo para formatear las columnas
    def __formatiar_columnas(self, df_filtrado):
        try:
            columnas_monedas, _ = self.__importar_json('columnas_monedas')
            columnas_texto, _ = self.__importar_json('columnas_texto')
            for col in columnas_monedas:
                if col in df_filtrado.columns:
                    # Asegurar que todos los valores sean string
                    df_filtrado[col] = df_filtrado[col].astype(str).str.strip()
                    # 🔹 Reemplazar símbolos de moneda (€, $, ₡, etc.) y espacios
                    df_filtrado[col] = df_filtrado[col].apply(lambda x: re.sub(r'[^\d,.-]', '', x) if isinstance(x, str) else x)
                    # 🔹 Manejar formato europeo "1.234,56" → "1234.56"
                    df_filtrado[col] = df_filtrado[col].apply(lambda x: x.replace('.', '') if x.count(',') == 1 else x)
                    df_filtrado[col] = df_filtrado[col].apply(lambda x: x.replace(',', '.') if x.count(',') == 1 else x)
                    # 🔹 Convertir a float
                    df_filtrado[col] = pd.to_numeric(df_filtrado[col], errors='coerce')
            # 🔹 Mantener columnas de identificación como texto
            for col in columnas_texto:
                if col in df_filtrado.columns:
                    df_filtrado[col] = df_filtrado[col].astype(str).str.strip()  # Asegurar que sea string

                    # 🔹 Verificar si el valor es numérico y convertirlo con separadores de miles
                    df_filtrado[col] = df_filtrado[col].apply(lambda x: "{:,.0f}".format(int(x)) if x.isdigit() else x)

                    # 🔹 Reemplazar comas por puntos (si es necesario)
                    df_filtrado[col] = df_filtrado[col].str.replace(',', '.')
            return df_filtrado
        except Exception as e:
            print(f"ERROR al formatear columnas: {e}")
            return None





    
    
    
    
    