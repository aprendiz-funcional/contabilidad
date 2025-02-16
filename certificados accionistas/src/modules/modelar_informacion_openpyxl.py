import openpyxl, os, json, random
import pandas as pd
from src.modules.leer_excel import LeerExcel  


from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl import load_workbook, Workbook

class   ModelarInformacion(LeerExcel):
    # Metodo constructor
    def __init__(self):
        path_ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.path_documentos = os.path.join(path_,"Documentos", "DIVIDENDOS.xlsx")
        #Ruta del Json
        self.__path_json = os.path.join(path_,'modules','Json','columnas_necesarias.json')
    # Metodo principal
    def main(self):
        try:
            # Obtener los datos depurados
            wb_datos = self.__depurar_informacion_dividendos()

            wb_retefuente = self.__extraer_retencion(wb_datos, '23651001', 'rete_fuente', 'retencion')

            wb_reteica = self.__extraer_retencion(wb_datos, '23680100', 'ret_eica', 'retencion')
            
            #wb_pagos_accionistas =self.__extraer_retencion(wb_datos, '23600501', 'pago_accionistas', 'columnas_permitidas')
            #wb_retefuente = self.__extraer_rete_fuente(wb_datos)

            #wb_reteica = self.__extraer_rete_ica(wb_datos)

            wb_pagos_accionistas = self.__extraer_pagos_accionistas(wb_datos, '23600501', 'pago_accionistas')

            
            wb_c_xp_contabilidad = self.__extraer_c_xp_contabilidad()

            wb_datos_cruzados = self.__cruzar_datos(wb_c_xp_contabilidad, wb_retefuente, wb_reteica, wb_pagos_accionistas)
            self.aplicar_estilos_excel(self.path_documentos)
            return wb_datos_cruzados


        except Exception as e:
            mensaje = f"ERROR al extraer retefuente: {e}"
            print(mensaje)
            return mensaje
    # Metodo para extraer los pagos de los accionistas
    def __extraer_pagos_accionistas(self, wb_datos, cuenta, tipo_retencion):
        try:
            # Filtrar datos por cuenta específica
            wb_datos = [fila for fila in wb_datos if fila.get("CUENTA") == cuenta]
            # 📌 Reestructurar los datos para que cada mes sea una columna
            # 📌 Reestructurar los datos para que cada mes sea una columna
            restructurado = {}
            for item in wb_datos:
                key = (item["TERCERO"], item["Nombre Tercero"])
                nombre_mes = item["Nombre Mes"].lower()  # Convertir el nombre del mes a minúsculas
                
                if key not in restructurado:
                    restructurado[key] = {"TERCERO": item["TERCERO"], "Nombre Tercero": item["Nombre Tercero"]}
                
                restructurado[key][nombre_mes] = item["DEBITOS"]

            # 📌 Convertir el diccionario en una lista ordenada
            bloques_meses = [
                (["enero", "febrero", "marzo", "abril", "mayo", "junio"], "Suma Enero-Junio"),
                (["julio", "agosto", "septiembre"], "Suma Julio-Septiembre"),
                (["octubre", "noviembre", "diciembre"], "Suma Octubre-Diciembre")
            ]
            
            lista_restructurada = []
            for key, valores in restructurado.items():
                fila = valores.copy()

                for meses, nombre_suma in bloques_meses:
                    fila[nombre_suma] = sum(fila.get(m, 0) for m in meses)
                
                lista_restructurada.append(fila)

            # 📌 Crear un nuevo archivo de Excel y escribir los datos
            nuevo_wb = Workbook()
            nuevo_ws = nuevo_wb.active

            # 📌 Escribir encabezados con sumas después de cada bloque de meses
            encabezados = ["TERCERO", "Nombre Tercero"]
            for meses, nombre_suma in bloques_meses:
                encabezados.extend(meses + [nombre_suma])

            nuevo_ws.append(encabezados)

            # 📌 Escribir filas con los datos
            for fila in lista_restructurada:
                nuevo_ws.append([fila.get(col, 0) for col in encabezados])

            # 📌 Aplicar formato a las columnas
            for col_num, col_name in enumerate(encabezados, start=1):
                col_letter = nuevo_ws.cell(row=1, column=col_num).column_letter
                nuevo_ws.column_dimensions[col_letter].width = 15  # Ajustar ancho de columnas
                
                for row_num in range(2, nuevo_ws.max_row + 1):
                    cell = nuevo_ws.cell(row=row_num, column=col_num)

                    # Formatear identificación con separador de miles
                    if col_name == "TERCERO":
                        cell.number_format = '#,##0'
                    
                    # Formatear valores monetarios en las columnas de meses y sumas
                    if col_name not in ["TERCERO", "Nombre Tercero"]:
                        cell.number_format = '"$"#,##0.00'
            # 📌 Convertir el diccionario en una lista ordenada
            columnas_permitidas = [
                "TERCERO","Nombre Tercero","enero", "febrero", "marzo", "abril", "mayo", "junio", "Suma Enero-Junio",
                "julio", "agosto", "septiembre", "Suma Julio-Septiembre",
                "octubre", "noviembre", "diciembre", "Suma Octubre-Diciembre"
            ]
            # 📌 Convertir nuevo_wb en una lista de diccionarios antes de retornarlo
            wb_datos = [
                {encabezados[i]: row[i] for i in range(len(encabezados))}
                for row in nuevo_ws.iter_rows(min_row=2, values_only=True)  # ⚠️ Se agrega values_only=True
            ]
            # Guardar los datos en el archivo de Excel
            self.__guardar_datos(wb_datos, columnas_permitidas, tipo_retencion)
            return wb_datos

        except Exception as e:
            mensaje = f"❌ ERROR al extraer pagos accionistas: {e}"
            print(mensaje)
            return mensaje



    def __extraer_pagos_accionistas1(self, wb_datos, cuenta):
        try:
            columnas_permitidas =  self.__importar_json('columnas_permitidas')
            print(f"📌 Columnas permitidas: {columnas_permitidas}")  # Debug

           

        except Exception as e:
            mensaje = f"ERROR al extraer retefuente: {e}"
            print(mensaje)
            return mensaje
    def __cruzar_datos(self, data_contabilidad, data_retefuente, data_reteica, data_pagos_accionistas):
        try:
            # Convertir listas de diccionarios a DataFrames, seleccionando solo las columnas necesarias
            df_contabilidad = pd.DataFrame(data_contabilidad)[["TERCERO", "Saldo Final"]]
            df_retefuente = pd.DataFrame(data_retefuente)[["TERCERO", "total_creditos"]]
            df_reteica = pd.DataFrame(data_reteica)[["TERCERO", "total_creditos"]]
            
            # En df_pagos_accionistas sí conservamos "Nombre Tercero"
            df_pagos_accionistas = pd.DataFrame(data_pagos_accionistas)

            # Renombrar columnas para evitar conflictos en la fusión
            df_retefuente.rename(columns={"CREDITOS": "RETEFUENTE"}, inplace=True)
            df_reteica.rename(columns={"CREDITOS": "RETEICA"}, inplace=True)

            # Fusionar los DataFrames en base a la columna "TERCERO"
            df_final = df_contabilidad.merge(df_retefuente, on="TERCERO", how="outer")
            df_final = df_final.merge(df_reteica, on="TERCERO", how="outer")
            df_final = df_final.merge(df_pagos_accionistas, on="TERCERO", how="outer")  # Aquí se mantiene "Nombre Tercero"

            # Crear nueva columna sumando los créditos de retefuente y reteica
            #df_final["CREDITOS_TOTALES"] = df_final[["CREDITOS_RETEFUENTE", "CREDITOS_RETEICA"]].sum(axis=1, min_count=1)

            # Reemplazar NaN con None para mantener el formato original
            df_final = df_final.where(pd.notna(df_final), None)
            encabezados = df_final.columns.tolist()
            # Convertir a lista de diccionarios garantizando el mismo formato de salida
            wb_final = df_final.to_dict(orient="records")
            self.__guardar_datos(wb_final, encabezados, "Datos Cruzados")
            return wb_final
        except Exception as e:
            mensaje = f"ERROR al cruzar datos: {e}"
            print(mensaje)
            return mensaje

    # Metodo para cruzar los datos  
    def __cruzar_datos1(self, data_contabilidad, data_retefuente, data_reteica, data_pagos_accionistas):
        try:
            # Convertir listas de diccionarios a DataFrames, seleccionando solo las columnas necesarias
            df_contabilidad = pd.DataFrame(data_contabilidad)[["TERCERO", "SALDO_FINAL"]]
            df_retefuente = pd.DataFrame(data_retefuente)[["TERCERO", "CREDITOS"]]
            df_reteica = pd.DataFrame(data_reteica)[["TERCERO", "CREDITOS"]]
            
            # En df_pagos_accionistas sí conservamos "Nombre Tercero"
            df_pagos_accionistas = pd.DataFrame(data_pagos_accionistas)

            # Nombre de la columna clave
            columna_clave = "TERCERO"

            # 📌 Seleccionar solo las columnas necesarias de cada DataFrame
            if " Saldo Final" in df_contabilidad.columns:
                df_contabilidad = df_contabilidad[[columna_clave, "SALDO FINAL"]]
            else:
                df_contabilidad["SALDO FINAL"] = 0  # Si no existe, crearla con valores None

            if "CREDITOS" in df_retefuente.columns:
                df_retefuente = df_retefuente[[columna_clave, "CREDITOS"]].rename(columns={"CREDITOS": "RETEFUENTE"})
            else:
                df_retefuente["RETEFUENTE"] = 0

            if "CREDITOS" in df_reteica.columns:
                df_reteica = df_reteica[[columna_clave, "CREDITOS"]].rename(columns={"CREDITOS": "RETEICA"})
            else:
                df_reteica["RETEICA"] = 0

            # 📌 Fusionar los DataFrames
            df_final = df_contabilidad.merge(df_retefuente, on=columna_clave, how="outer")
            df_final = df_final.merge(df_reteica, on=columna_clave, how="outer")
            df_final = df_final.merge(df_pagos_accionistas, on=columna_clave, how="outer")

            # 📌 Agregar las columnas que se extrajeron previamente (opcional, debes definir cuáles)
            #df_final["CREDITOS_TOTALES"] = df_final[["CREDITOS_RETEFUENTE", "CREDITOS_RETEICA"]].sum(axis=1, min_count=1)

            # 📌 Reemplazar NaN con None
            df_final = df_final.where(pd.notna(df_final), 0)

            # 📌 Convertir a lista de diccionarios y retornar
            return df_final.to_dict(orient="records")
        except Exception as e:
            mensaje = f"ERROR al cruzar datos: {e}"
            print(mensaje)
            return mensaje

    

    def __cruzar_datos1(self, data_contabilidad, data_retefuente, data_reteica, data_pagos_accionistas):
        # Convertir listas de diccionarios a DataFrames
        df_contabilidad = pd.DataFrame(data_contabilidad)
        df_retefuente = pd.DataFrame(data_retefuente)
        df_reteica = pd.DataFrame(data_reteica)
        df_pagos_accionistas = pd.DataFrame(data_pagos_accionistas)

        # Nombre de la columna clave
        columna_clave = "TERCERO"

        # Fusionar los DataFrames en base a la columna "Tercero"
        df_final = df_contabilidad.merge(df_retefuente, on=columna_clave, how="outer", suffixes=("_contabilidad", "_retefuente"))
        df_final = df_final.merge(df_reteica, on=columna_clave, how="outer", suffixes=("", "_reteica"))
        df_final = df_final.merge(df_pagos_accionistas, on=columna_clave, how="outer", suffixes=("", "_pagos_accionistas"))

        # Reemplazar NaN con None para mantener el formato original
        df_final = df_final.where(pd.notna(df_final), None)

        # Lista de columnas que deben mantener formato de moneda o número con separador decimal
        columnas_monedas = ["Valor", "Monto", "Total", "Impuesto"]  # Ajusta con los nombres correctos de tus columnas

        # Aplicar formato de moneda o número a las columnas necesarias
        for col in df_final.columns:
            if col in columnas_monedas and df_final[col].dtype in [float, int]:
                df_final[col] = df_final[col].apply(lambda x: f"{x:,.2f}" if x is not None else None)

        # Convertir a lista de diccionarios garantizando el mismo formato de salida
        return df_final.to_dict(orient="records")

    def __cruzar_datos1(self, wb_c_xp_contabilidad, wb_retefuente, wb_reteica, wb_pagos_accionistas):
        pass
    
    # Metodo para extraer los datos de c_xp_contabilidad
    def __extraer_c_xp_contabilidad(self):
        try:
            wb_c_xp_contabilidad, mensaje = super().leer_datos_excel(self.path_documentos, ' c xp contabilidad')
            
            if wb_c_xp_contabilidad is None:
                print(f"📌 ERROR: {mensaje}")
                return mensaje

            print(f"📌 Tipo de datos recibidos: {type(wb_c_xp_contabilidad)}")  
            print(f"📌 Primeros datos recibidos: {wb_c_xp_contabilidad[:3] if wb_c_xp_contabilidad else 'No hay datos'}")


            # Normalizar nombres de columnas en los datos (evitar problemas con espacios extras)
            for fila in wb_c_xp_contabilidad:
                for key in list(fila.keys()):  # Iterar sobre una copia de las claves
                    fila[key.strip()] = fila.pop(key)  


            return wb_c_xp_contabilidad

        except Exception as e:
            mensaje = f"ERROR al depurar los datos de excel<br> {e}"
            print(mensaje)
            return mensaje
    # Metodo para extraer la rete ica
    def __extraer_rete_ica(self, wb_datos):
        try:
            # Filtrar por la cuenta 1239852 y extraer columnas específicas
            # Cargar columnas permitidas desde el JSON
            columnas_permitidas, mensaje = self.__importar_json('retetencion')
            print(f"📌 Columnas permitidas: {columnas_permitidas}")  # Debug

            if not isinstance(columnas_permitidas, list):
                mensaje =f"ERROR: 'columnas_permitidas' debe ser una lista, pero se recibió {type(columnas_permitidas)}"
                print(mensaje)
                return mensaje
            
            datos_filtrados = [
                {columna: fila.get(columna) for columna in columnas_permitidas}
                for fila in wb_datos if fila.get("CUENTA") == "23651001"
            ]

            # Ordenar los datos de menor a mayor por la columna TERCERO
            datos_filtrados = sorted(datos_filtrados, key=lambda x: x["TERCERO"])

            
            if not datos_filtrados:
                return "No se encontraron registros con la cuenta 23651001."
            
            self.__guardar_datos(datos_filtrados, columnas_permitidas, "ReteFuente_sin filtar")
            datos_filtrados = self.__agrupar_por_identificacion(datos_filtrados)
            return datos_filtrados

        except Exception as e:
            mensaje = f"ERROR al extraer retefuente: {e}"
            print(mensaje)
            return mensaje
    # Metodo para extraer la rete fuente
    def __extraer_rete_fuente(self, wb_datos):
        try:
            # Filtrar por la cuenta 1239852 y extraer columnas específicas
            # Cargar columnas permitidas desde el JSON
            columnas_permitidas, mensaje = self.__importar_json('retencion')
            print(f"📌 Columnas permitidas: {columnas_permitidas}")  # Debug

            if not isinstance(columnas_permitidas, list):
                mensaje =f"ERROR: 'columnas_permitidas' debe ser una lista, pero se recibió {type(columnas_permitidas)}"
                print(mensaje)
                return mensaje
            
            datos_filtrados = [
                {columna: fila.get(columna) for columna in columnas_permitidas}
                for fila in wb_datos if fila.get("CUENTA") == "23651001"
            ]

            # Ordenar los datos de menor a mayor por la columna TERCERO
            datos_filtrados = sorted(datos_filtrados, key=lambda x: x["TERCERO"])

            
            if not datos_filtrados:
                return "No se encontraron registros con la cuenta 23651001."
            
            self.__guardar_datos(datos_filtrados, columnas_permitidas, "ReteFuente_sin filtar")
            datos_filtrados = self.__agrupar_por_identificacion(datos_filtrados)
            return datos_filtrados

        except Exception as e:
            mensaje = f"ERROR al extraer retefuente: {e}"
            print(mensaje)
            return mensaje
    # Metodo para depurar la informacion
    def __depurar_informacion_dividendos(self):
        try:
            datos, mensaje = super().leer_datos_excel(self.path_documentos, 'DIVIDENDOS')
            
            if datos is None:
                print(f"📌 ERROR: {mensaje}")
                return mensaje

            print(f"📌 Tipo de datos recibidos: {type(datos)}")  
            print(f"📌 Primeros datos recibidos: {datos[:3] if datos else 'No hay datos'}")

            # Cargar columnas permitidas desde el JSON
            columnas_permitidas, mensaje = self.__importar_json('columnas_permitidas')
            print(f"📌 Columnas permitidas: {columnas_permitidas}")  # Debug

            if not isinstance(columnas_permitidas, list):
                mensaje =f"ERROR: 'columnas_permitidas' debe ser una lista, pero se recibió {type(columnas_permitidas)}"
                print(mensaje)
                return mensaje

            # Normalizar nombres de columnas en los datos (evitar problemas con espacios extras)
            for fila in datos:
                for key in list(fila.keys()):  # Iterar sobre una copia de las claves
                    fila[key.strip()] = fila.pop(key)  

            # Filtrar solo las columnas necesarias
            datos_filtrados = [
                {col: fila[col] for col in columnas_permitidas if col in fila}
                for fila in datos
            ]

            # Guardar en Excel
            self.__guardar_datos(datos_filtrados, columnas_permitidas, "Datos Filtrados")

            return datos_filtrados

        except Exception as e:
            mensaje = f"ERROR al depurar los datos de excel<br> {e}"
            print(mensaje)
            return mensaje
    #Metodo para tomar el diccionario del json
    def __importar_json(self, nom_json):
        try:                    
            with open(self.__path_json, "r", encoding="utf-8") as file:
                users = json.load(file)
            # Acceder a los datos
            self.campos_formulario = users.get(nom_json, {})
            mensaje = f"Json importado correctamente {nom_json}"
            print(mensaje)
            return self.campos_formulario, mensaje
        except Exception as e:
            mensaje = f"E importando el Json: del formulario {nom_json}"
            print(f"{mensaje} {e}")
            return None, mensaje
    # Metodo para agrupar los datos por identificacion
    def __agrupar_por_identificacion(self, wb_datos, tipo_retencion):
        try:
            agrupados = {}
            for fila in wb_datos:
                identificacion = fila.get("TERCERO")
                nombre_tercero = fila.get("Nombre Tercero", "Desconocido")  # Si no hay nombre, asigna "Desconocido"
                
                # Asegurar que CREDITOS es un número, eliminando espacios y reemplazando comas
                creditos = fila.get("CREDITOS", 0)
                try:
                    creditos = float(str(creditos).replace(",", "").strip())
                except ValueError:
                    print(f"⚠️ Valor no numérico en CREDITOS: {creditos}")
                    creditos = 0  # Si hay error en la conversión, asignar 0
                
                # Debug: imprimir valores antes de agrupar
                print(f"📊 Procesando: TERCERO={identificacion}, CREDITOS={creditos}")

                # Si la identificación no está en el diccionario, inicializarla
                if identificacion not in agrupados:
                    agrupados[identificacion] = {
                        "TERCERO": identificacion, 
                        "Nombre Tercero": nombre_tercero, 
                        "total_creditos": 0  # Inicializar el total de créditos
                    }
                
                # Sumar los créditos al identificador correspondiente
                agrupados[identificacion]["total_creditos"] += creditos
            
            # Convertir el diccionario en una lista de valores
            resultado = list(agrupados.values())
            nombre_hoja = f'Agrupados  {tipo_retencion}'
            # Guardar los datos agrupados en el sistema
            self.__guardar_datos(resultado, ["TERCERO", "Nombre Tercero", "total_creditos"], nombre_hoja)
            
            return resultado

        except Exception as e:
            # Manejo de errores
            mensaje = f"ERROR al agrupar datos por identificación: {e}"
            print(mensaje)
            return mensaje
    # Metodo para guardar los datos
    def __guardar_datos(self, wb_datos, columnas_permitidas, nomhoja):
        try:
            # Cargar el archivo de Excel
            wb = openpyxl.load_workbook(self.path_documentos)

            # Verificar si la hoja ya existe
            if nomhoja in wb.sheetnames:
                ws = wb[nomhoja]
            else:
                ws = wb.create_sheet(title=nomhoja)

            # Si la hoja está vacía, agregar encabezados
            if ws.max_row == 1:
                ws.append(columnas_permitidas)

            # Escribir los datos en la hoja
            for fila in wb_datos:
                ws.append([fila.get(col, "") for col in columnas_permitidas])  

            # Guardar el archivo sin sobrescribir otras hojas
            wb.save(self.path_documentos)
            wb.close()  # Cerrar el archivo para evitar bloqueos

        except Exception as e:
            mensaje = f"ERROR al guardar los datos: {e}"
            print(mensaje)
            return mensaje


    # Metodo para extraer la rete fuente
    def __extraer_retencion(self, wb_datos, cuenta, tipo_retencion, json):
        try:
            # Filtrar por la cuenta 1239852 y extraer columnas específicas
            # Cargar columnas permitidas desde el JSON
            columnas_permitidas, mensaje = self.__importar_json(json)
            print(f"📌 Columnas permitidas: {columnas_permitidas}")  # Debug

            if not isinstance(columnas_permitidas, list):
                mensaje =f"ERROR: 'columnas_permitidas' debe ser una lista, pero se recibió {type(columnas_permitidas)}"
                print(mensaje)
                return mensaje
            
            datos_filtrados = [
                {columna: fila.get(columna) for columna in columnas_permitidas}
                for fila in wb_datos if fila.get("CUENTA") == cuenta
            ]

            # Ordenar los datos de menor a mayor por la columna TERCERO
            datos_filtrados = sorted(datos_filtrados, key=lambda x: x["TERCERO"])

            
            if not datos_filtrados:
                return None, f"No se encontraron registros con la cuenta {cuenta}."
            
            #self.__guardar_datos(datos_filtrados, columnas_permitidas, tipo_retencion)
            datos_filtrados = self.__agrupar_por_identificacion(datos_filtrados, tipo_retencion)
            return datos_filtrados

        except Exception as e:
            mensaje = f"ERROR al extraer retefuente: {e}"
            print(mensaje)
            return mensaje
    
    #Metodo para crear las carpetas
    def crear_carpeta(self, nom_carpeta):
            try:
                if not os.path.exists(self.path_historicos):
                    os.mkdir(self.path_historicos)
                self.carpeta_nomPaso = os.path.join(self.path_historicos, nom_carpeta)
                if not os.path.exists(self.carpeta_nomPaso):
                    os.mkdir(self.carpeta_nomPaso)
                self.carpeta_mes = os.path.join(self.path_historicos, nom_carpeta, dt.datetime.today().strftime("%Y-%B-%d"))
                if not os.path.exists(self.carpeta_mes):
                    os.mkdir(self.carpeta_mes)
                #ruta para la creacion del logger
                log_file = os.path.join(self.carpeta_mes, f"log_{dt.datetime.now().strftime(f'{nom_carpeta}-%d-%m-%Y')}.txt")
                return log_file
            except Exception as e:
                self.logger.log(f"ERROR creando las carpetas: {e}<br>")
                return None

    #Metodo para aplicar estilos a los archivos de excel
    def aplicar_estilos_excel(self, ruta_excel):
        try:
            # Cargar el archivo existente
            wb = load_workbook(ruta_excel)
            # Definir estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            alignment_encabezado = Alignment(horizontal="center", vertical="center")
            alignment_cuerpo = Alignment(horizontal="left", vertical="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            # Generar colores para las pestañas de las hojas
            def generar_color_aleatorio():
                return f"{random.randint(0, 255):02X}{random.randint(0, 255):02X}{random.randint(0, 255):02X}"
            # Iterar por cada hoja del archivo
            for ws in wb.worksheets:
                # Aplicar estilos a los encabezados (primera fila)
                for cell in ws[1]:  # Supone que la primera fila son los encabezados
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment_encabezado
                    cell.border = thin_border
                # Aplicar bordes y alineación a todas las celdas con datos
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.alignment = alignment_cuerpo
                        cell.border = thin_border
                # Ajustar el ancho de las columnas automáticamente
                for col in ws.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = max_length + 2
                # Asignar color aleatorio a la pestaña de la hoja
                ws.sheet_properties.tabColor = generar_color_aleatorio()
            # Guardar los cambios
            wb.save(ruta_excel)
            print("Estilos aplicados y colores asignados correctamente a las hojas del archivo de Excel.")
        except Exception as e:
            print(f"ERROR: No se pudieron aplicar estilos y colores al archivo de Excel: {str(e)}")