import openpyxl

class LeerExcel:
    # metodo constructor
    def __init__(self):
        pass
    # Leer datos de un archivo excel
    def leer_datos_excel(self, ruta, nomhoja):
        try:
           wb = openpyxl.load_workbook(ruta)
           ws = wb[nomhoja]
           # Obtener los encabezados de la primera fila
           encabezados = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
           # Convertir las filas en diccionarios
           datos = [
                {encabezados[i]: cell.value for i, cell in enumerate(row)}
                for row in ws.iter_rows(min_row=2)
            ]
           mensaje = f"Datos de excel leidos con exito en la {nomhoja}"
           return datos, mensaje
        except Exception as e:
            mensaje = f"ERROR al leer los datos de excel<br>"
            print(f"{mensaje} {e}")
            return None, mensaje